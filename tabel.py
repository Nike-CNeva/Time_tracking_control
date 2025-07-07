import os
import sys
import sqlite3
import holidays
import holidays.countries
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from ctypes import *
import logging
from PyQt5.QtCore import QTimer, QTime, QByteArray, Qt, QDate, QBuffer, QIODevice, QObject, QThread, pyqtSignal
from PyQt5.QtGui import QColor, QTransform, QPixmap
from PyQt5.QtWidgets import QCheckBox, QComboBox, QTableWidget, QTableWidgetItem, QCalendarWidget, QLabel, QMessageBox, QDateEdit, QLineEdit, QApplication, QDialog, QWidget, QListWidget, QListWidgetItem, QMainWindow, QSlider, QHBoxLayout, QGraphicsView, QGraphicsScene, QGraphicsPixmapItem, QGraphicsRectItem, QVBoxLayout, QPushButton, QFileDialog, QRadioButton, QProgressBar, QTextEdit
# Определяем путь к базе данных и логу

DB_PATH = os.path.join(r"\\10.42.0.112\TimeTrackingSystem", "db", "employee_management.db")
LOG_PATH = os.path.join(r"C:\TimeTrackingSystem", "logs", "timesheet_log.txt")
DLL_PATH = os.path.join(r"C:\TimeTrackingSystem", "libs", "ftrapi.dll")
# Настроим логирование
logging.basicConfig(
    filename=LOG_PATH,  # Файл для записи логов
    level=logging.INFO,  # Уровень логирования
    format='%(asctime)s - %(message)s',  # Формат записи
)

FTR_RETCODE_OK                  = 0
FTR_PARAM_IMAGE_WIDTH           = c_ulong(1)
FTR_PARAM_IMAGE_HEIGHT          = c_ulong(2)
FTR_PARAM_IMAGE_SIZE            = c_ulong(3)
FTR_PARAM_CB_FRAME_SOURCE       = c_ulong(4)
FTR_PARAM_CB_CONTROL            = c_ulong(5)
FTR_PARAM_MAX_TEMPLATE_SIZE     = c_ulong(6)
FTR_PARAM_MAX_FAR_REQUESTED     = c_ulong(7)
FTR_PARAM_SYS_ERROR_CODE        = c_ulong(8)
FTR_PARAM_FAKE_DETECT           = c_ulong(9)
FTR_PARAM_MAX_MODELS            = c_ulong(10)
FTR_PARAM_FFD_CONTROL           = c_ulong(11)
FTR_PARAM_MIOT_CONTROL          = c_ulong(12)
FTR_PARAM_MAX_FARN_REQUESTED    = c_ulong(13)
FTR_PARAM_VERSION               = c_ulong(14)

FSD_FUTRONIC_USB    = c_void_p(1)

FTR_CB_RESP_CANCEL      = c_ulong(1)
FTR_CB_RESP_CONTINUE    = c_ulong(2)

FTR_PURPOSE_IDENTIFY = c_ulong(2)
FTR_PURPOSE_ENROLL   = c_ulong(3)
FTR_PURPOSE_COMPATIBILITY = c_ulong(4)

FTR_STATE_FRAME_PROVIDED    = 0x01
FTR_STATE_SIGNAL_PROVIDED   = 0x02

FTR_VERSION_CURRENT = c_ulong(3)

# Определение типов
FTR_USER_CTX = c_void_p  # Аналог указателя на произвольные данные
FTR_STATE = c_uint32     # Маска состояния
FTR_SIGNAL = c_uint32    # Тип сигнала
FTR_BITMAP_PTR = c_void_p  # Указатель на bitmap-структуру
FTR_RESPONSE = c_ulong   # Аналог typedef UDGT32 FTR_RESPONSE

class EmployeeManager:
    def __init__(self):
        self.db_connection = self.connect_to_database()
    
    def connect_to_database(self):
        try:
            connection = sqlite3.connect(DB_PATH)
            return connection
        except sqlite3.Error as err:
            logging.error(f"Ошибка подключения к базе данных: {err}")
            raise

    def get_all_employees(self):
        try:
            cursor = self.db_connection.cursor()
            cursor.execute("SELECT id, lastname, firstname, patronymic, dob, status, department, wages FROM employees ORDER BY CASE WHEN department = 'Руководство' THEN 1 WHEN department = 'Смена 1' THEN 2 WHEN department = 'Смена 2' THEN 3 WHEN department = 'Прокат' THEN 4 WHEN department = 'Резка' THEN 5 WHEN department = 'Склад' THEN 6 ELSE 7 END, lastname, firstname;")
            result = cursor.fetchall()
            return [{'id': emp_id, 'lastname': lastname, 'firstname': firstname, 'patronymic': patronymic, 'dob': dob,
                'status': status, 'department': department, 'wages' : wages} 
                    for emp_id, lastname, firstname, patronymic, dob, status, department, wages in result]
        except sqlite3.Error as err:
            logging.error(f"Ошибка при получении списка сотрудников: {err}")
            return []
        finally:
            cursor.close()

    def has_timesheet_records(self, employee_id, year_month):
        """Проверить, есть ли у сотрудника записи за указанный месяц."""
        try:
            cursor = self.db_connection.cursor()
            cursor.execute("""
                SELECT COUNT(*) 
                FROM timesheet 
                WHERE employee_id = ? AND strftime('%Y-%m', date) = ?
            """, (employee_id, year_month))
            result = cursor.fetchone()
            return result[0] > 0
        except sqlite3.Error as err:
            logging.error(f"Ошибка при проверке записей табеля: {err}")
            return False
        finally:
            cursor.close()

    def get_dismissed_employees(self):
        """Получить список уволенных сотрудников."""
        try:
            cursor = self.db_connection.cursor()
            cursor.execute("SELECT id, lastname, firstname, patronymic FROM employees WHERE status = 'уволен'")
            result = cursor.fetchall()
            return [{'id': emp_id, 'lastname': lastname, 'firstname': firstname, 'patronymic': patronymic} 
                    for emp_id, lastname, firstname, patronymic in result]
        except sqlite3.Error as err:
            logging.error(f"Ошибка при получении списка уволенных сотрудников: {err}")
            return []
        finally:
            cursor.close()

    def add_employee(self, lastname, firstname, patronymic, dob, hire_date, position):
        try:
            cursor = self.db_connection.cursor()
            cursor.execute("""
                INSERT INTO employees (lastname, firstname, patronymic, dob, hire_date, position)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (lastname, firstname, patronymic, dob, hire_date, position))
            self.db_connection.commit()
        except sqlite3.Error as err:
            logging.error(f"Ошибка при добавлении нового сотрудника: {err}")
            return []
        finally:
            cursor.close()

    def get_employee_by_id(self, emp_id):
        try:
            cursor = self.db_connection.cursor()
            cursor.execute(
                "SELECT id, lastname, firstname, patronymic, dob, hire_date, position, photo, status, department, wages FROM employees WHERE id = ?",
                (emp_id,)
            )
            result = cursor.fetchone()
            if result:
                return {
                    'id': result[0],
                    'lastname': result[1],
                    'firstname': result[2],
                    'patronymic': result[3],
                    'dob': result[4],
                    'hire_date': result[5],
                    'position': result[6],
                    'photo': result[7],
                    'status': result[8],
                    'department': result[9],
                    'wages': result[10]
                }
            return None
        except sqlite3.Error as err:
            logging.error(f"Ошибка получения данных о сотруднике: {err}")
            return []
        finally:
            cursor.close()

    def get_employee_photo(self, emp_id):
        try:
            cursor = self.db_connection.cursor()
            cursor.execute("SELECT photo FROM employees WHERE id = ?", (emp_id,))
            result = cursor.fetchone()
            return result[0] if result else None
        except sqlite3.Error as err:
            logging.error(f"Ошибка при получении фото сотрудника: {err}")
            return []
        finally:
            cursor.close()

    def update_employee(self, emp_id, lastname, firstname, patronymic, dob, hire_date, position, wages, status, department):
        try:
            cursor = self.db_connection.cursor()
            cursor.execute("""
                UPDATE employees
                SET lastname = ?, firstname = ?, patronymic = ?, dob = ?, hire_date = ?, position = ?, wages = ?, status = ?, department = ?
                WHERE id = ?
            """, (lastname, firstname, patronymic, dob, hire_date, position, wages, status, department, emp_id))
            self.db_connection.commit()
        except sqlite3.Error as err:
            logging.error(f"Ошибка обновления данных сотрудника: {err}")
            return []
        finally:
            cursor.close()

    def delete_employee(self, emp_id):
        try:
            cursor = self.db_connection.cursor()
            cursor.execute("DELETE FROM employees WHERE id = ?", (emp_id,))
            self.db_connection.commit()
        except sqlite3.Error as err:
            logging.error(f"Ошибка при удалении сотрудника: {err}")
            return []
        finally:
            cursor.close()

    def get_timesheet(self, year, month):
        try:
            cursor = self.db_connection.cursor()
            query = """SELECT employee_id, date, arrival_time, departure_time 
                    FROM timesheet WHERE strftime('%Y', date) = ? AND strftime('%m', date) = ?"""

            cursor.execute(query, (str(year), str(month).zfill(2)))
            
            # Получаем имена столбцов
            column_names = [description[0] for description in cursor.description]
            
            # Преобразуем кортежи в словари
            result = cursor.fetchall()

            return [dict(zip(column_names, row)) for row in result]
            
        except sqlite3.Error as err:
            logging.error(f"Ошибка при получении данных табеля: {err}")
            return []
        finally:
            cursor.close()

    def get_timesheet_entry(self, employee_id, date):
        """Получить запись табеля для конкретного сотрудника и даты."""
        try:
            cursor = self.db_connection.cursor()    
            query = """
                SELECT arrival_time, departure_time
                FROM timesheet
                WHERE employee_id = ? AND date = ?
            """
            cursor.execute(query, (employee_id, date))
            
            # Получаем имена столбцов
            column_names = [description[0] for description in cursor.description]
            
            result = cursor.fetchone()
            if result:
                return dict(zip(column_names, result))
            
            return None
        except sqlite3.Error as err:
            logging.error(f"Ошибка при получении записи табеля для сотрудника {employee_id} на дату {date}: {err}")
            return None
        finally:
            cursor.close()

    def update_timesheet(self, employee_id, date, arrival_time=None, departure_time=None):
        try:
            cursor = self.db_connection.cursor()    
            query = """
                INSERT INTO timesheet (employee_id, date, arrival_time, departure_time)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(employee_id, date) DO UPDATE SET
                    arrival_time = COALESCE(EXCLUDED.arrival_time, arrival_time),
                    departure_time = COALESCE(EXCLUDED.departure_time, departure_time)
            """
            cursor.execute(query, (employee_id, date, arrival_time, departure_time))
            self.db_connection.commit()
        except sqlite3.Error as err:
            logging.error(f"Ошибка при обновлении данных: {err}")
            self.db_connection.rollback()
        finally:
            cursor.close()

    def remove_timesheet(self, employee_id, date):
        """Удаление записи о времени для сотрудника и дня."""
        try:
            cursor = self.db_connection.cursor()
            query = "DELETE FROM timesheet WHERE employee_id = ? AND date = ?"
            cursor.execute(query, (employee_id, date))
            self.db_connection.commit()  # Подтверждение изменений
        except sqlite3.Error as err:
            logging.error(f"Ошибка при удалении данных: {err}")
            self.db_connection.rollback()  # Откат изменений в случае ошибки
        finally:
            cursor.close()

    
    def add_or_update_fingerprint(self, employee_id, finger_name, fingerprint_template, quality):
        #Добавление или обновление отпечатка пальца.
        try:
                cursor = self.db_connection.cursor()
                # Проверяем, есть ли уже отпечаток для этого пальца
                cursor.execute("""
                    SELECT id FROM fingerprints 
                    WHERE employee_id = ? AND finger_name = ?
                """, (employee_id, finger_name))
                result = cursor.fetchone()

                if result:
                    # Обновляем существующий отпечаток
                    cursor.execute("""
                        UPDATE fingerprints 
                        SET fingerprint_template = ?, quality = ?
                        WHERE id = ?
                    """, (fingerprint_template, quality, result[0]))
                else:
                    # Добавляем новый отпечаток
                    cursor.execute("""
                        INSERT INTO fingerprints (employee_id, finger_name, fingerprint_template, quality)
                        VALUES (?, ?, ?, ?)
                    """, (employee_id, finger_name, fingerprint_template, quality))
                self.db_connection.commit()
        except sqlite3.Error as err:
            logging.error(f"Ошибка при сохранении отпечатка: {err}")
            self.db_connection.rollback()
        finally:
            cursor.close()
    
    def get_templates_from_database(self):
        try:
                cursor = self.db_connection.cursor()
                cursor.execute("SELECT employee_id, fingerprint_template FROM fingerprints")
                result = cursor.fetchall()
                return [(row[0], row[1]) for row in result]
        except sqlite3.Error as err:
            logging.error(f"Ошибка при получении шаблонов: {err}")
            return []
        finally:
            cursor.close()
    
    def add_leave(self, employee_id, start_date, end_date, leave_type):
        try:
            cursor = self.db_connection.cursor()
            cursor.execute("""
                INSERT INTO leaves (employee_id, start_date, end_date, type)
                VALUES (?, ?, ?, ?)
            """, (employee_id, start_date, end_date, leave_type))
            self.db_connection.commit()
        except sqlite3.Error as err:
            logging.error(f"Ошибка при добавлении отпуска/больничного: {err}")
            self.db_connection.rollback()
        finally:
            cursor.close()

    def get_leaves_for_employee(self, employee_id, year):
        try:
            cursor = self.db_connection.cursor()
            cursor.execute("""
                SELECT type, start_date, end_date
                FROM leaves
                WHERE employee_id = ? AND strftime('%Y', start_date) = ?
            """, (employee_id, str(year)))
            return cursor.fetchall()
        except sqlite3.Error as err:
            logging.error(f"Ошибка при получении данных отпусков: {err}")
            return []
        finally:
            cursor.close()

    def get_leave_days_for_year(self, employee_id, year):
        try:
            cursor = self.db_connection.cursor()
            query = """
                SELECT type, start_date, end_date
                FROM leaves
                WHERE employee_id = ? AND (
                    strftime('%Y', start_date) = ? OR strftime('%Y', end_date) = ?
                )
            """
            cursor.execute(query, (employee_id, str(year), str(year)))
            leaves = cursor.fetchall()
            
            total_leave_days = {"Отпуск": 0, "Больничный": 0}
            current_date = QDate.currentDate().toPyDate()

            for leave_type, start_date, end_date in leaves:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                
                # Учитываем только дни до текущей даты включительно
                effective_end_date = min(end_date, current_date)
                
                if start_date <= effective_end_date:
                    total_leave_days[leave_type] += (effective_end_date - start_date).days + 1

            return total_leave_days
        except sqlite3.Error as err:
            logging.error(f"Ошибка при расчете дней: {err}")
            return {"Отпуск": 0, "Больничный": 0}
        finally:
            cursor.close()

    def calculate_salary(self, employee_id, year, month):
        try:
            cursor = self.db_connection.cursor()
            query = """
                SELECT arrival_time, departure_time
                FROM timesheet
                WHERE employee_id = ? AND strftime('%Y', date) = ? AND strftime('%m', date) = ?
            """
            cursor.execute(query, (employee_id, str(year), str(month).zfill(2)))
            records = cursor.fetchall()

            total_hours = 0
            days = 0
            for arrival_time, departure_time in records:
                if arrival_time and departure_time:
                    arrival = QTime.fromString(arrival_time, "HH:mm")
                    departure = QTime.fromString(departure_time, "HH:mm")
                    
                    if arrival.isValid() and departure.isValid():
                        worked_minutes = arrival.msecsTo(departure) / 60000
                        if worked_minutes > 240:
                            worked_minutes -= 30  # Учитываем обед
                        # Округляем время только при экспорте
                        rounded_minutes = round(worked_minutes / 30) * 30
                        worked_hours = rounded_minutes /60
                        total_hours += worked_hours
                        days += 1

            # Получаем ставку сотрудника
            cursor.execute("SELECT wages FROM employees WHERE id = ?", (employee_id,))
            result = cursor.fetchone()
            wages = int(result[0]) if result and result[0] is not None else 0

            if wages > 5000:
                # Установим список праздников для России
                year = datetime.today().year
                month = datetime.today().month
                russian_holidays = holidays.RU(year)
                current_date = datetime(year, month, 1)
                last_day_of_month = datetime(year, month + 1, 1) - timedelta(days=1)
                working_days = 0
                while current_date <= last_day_of_month:
                    if current_date.weekday() < 5 and current_date not in russian_holidays:
                        working_days += 1
                    current_date += timedelta(days=1)
                # Рассчитываем ставку в час по окладу
                hourly_wage = wages / working_days
                # Расчитываем зарплату
                salary = days * hourly_wage
            else:
                # Если ставка меньше или равна 5000, расчет по обычной логике
                salary = total_hours * wages

            return salary
        except sqlite3.Error as err:
            logging.error(f"Ошибка при расчете зарплаты: {err}")
            return 0
        finally:
            cursor.close()

class TimesheetWindow(QDialog):
    def __init__(self, employee_manager, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Табель учета рабочего времени")
        self.setGeometry(100, 100, 800, 600)
        self.employee_manager = employee_manager

        year = datetime.today().year
        self.holidays = holidays.RU(year)

        # Создание выпадающего списка для выбора группы подразделений
        self.group_selector = QComboBox()
        self.group_selector.addItems(["Все", "Руководство", "Производство", "Склад"])
        self.group_selector.currentTextChanged.connect(self.load_timesheet) 

        self.calendar = QCalendarWidget()
        self.calendar.setGridVisible(True)
        self.calendar.selectionChanged.connect(self.load_timesheet)
        self.calendar.paintCell = self.paint_cell

        self.table = QTableWidget()

        self.load_button = QPushButton("Обновить табель")
        self.load_button.clicked.connect(self.load_timesheet)

        self.save_button = QPushButton("Ввести данные")
        self.save_button.clicked.connect(self.add_timesheet_entry)

        self.export_button = QPushButton("Экспорт в Excel")
        self.export_button.clicked.connect(self.export_to_excel)

        layout = QVBoxLayout()
        layout.addWidget(self.group_selector)
        layout.addWidget(self.calendar)
        layout.addWidget(self.table)
        layout.addWidget(self.load_button)
        layout.addWidget(self.save_button)
        layout.addWidget(self.export_button)

        self.setLayout(layout)

        self.load_timesheet()

    def on_date_selected(self):
        """Обработчик для изменения выбранной даты в календаре"""
        self.calendar.update()  # Перерисовываем календарь

    def paint_cell(self, painter, rect, date):
        # Если дата является праздничной, меняем её стиль
        date = date.toPyDate()
        # Получаем выбранную дату из календаря
        selected_date = self.calendar.selectedDate().toPyDate()
        if date == selected_date:
            painter.fillRect(rect, QColor(173, 216, 230))  # Светло-голубой фон для выбранного дня
            painter.setPen(QColor(0, 0, 0))  # Черный текст для выбранного дня
            painter.drawText(rect, Qt.AlignCenter, str(date.day))
        elif date in self.holidays:
            painter.fillRect(rect, QColor(255, 255, 255))  # Белый фон
            painter.setPen(QColor(255, 0, 0))  # Красный текст
            painter.drawText(rect, Qt.AlignCenter, str(date.day))
        elif date.weekday() == 5 or date.weekday() == 6:  # 5 - суббота, 6 - воскресенье
            painter.fillRect(rect, QColor(255, 255, 255))  # Белый фон для выходных
            painter.setPen(QColor(255, 0, 0))  # Красный текст для выходных
            painter.drawText(rect, Qt.AlignCenter, str(date.day))  # День месяца
        else:
            # Стандартный рендеринг ячейки для обычных дней
            painter.setPen(QColor(0, 0, 0))  # Черный цвет для обычных дней
            painter.drawText(rect, Qt.AlignCenter, str(date.day))

    def timedelta_to_time_str(self, arrival_time, departure_time):
        """Преобразует строки времени в объект QTime для работы с ними."""

        arrival_time_str = QTime.fromString(arrival_time, "HH:mm")
        departure_time_str = QTime.fromString(departure_time, "HH:mm")

        return arrival_time_str, departure_time_str

    def get_leave_status(self, employee_id, year, month, day):
        """Проверяет, находится ли сотрудник на больничном или в отпуске в заданную дату."""
        leaves = self.employee_manager.get_leaves_for_employee(employee_id, year)
        date = QDate(year, month, day)
        for leave in leaves:
            start_date = QDate.fromString(leave[1], "yyyy-MM-dd")
            end_date = QDate.fromString(leave[2], "yyyy-MM-dd")
            if start_date <= date <= end_date:
                return "Больничный" if leave[0] == "Больничный" else "Отпуск"
        return None

    def load_timesheet(self):
        # Определение текущей выбранной группы
        selected_group = self.group_selector.currentText()

        # Получение выбранной даты
        date = self.calendar.selectedDate()
        year, month = date.year(), date.month()
        # Форматируем текущий год и месяц для SQL-запроса
        current_year_month = f"{year}-{str(month).zfill(2)}"
        # Получаем количество дней в выбранном месяце
        days_in_month = date.daysInMonth()
        
        # Определяем группы подразделений
        groups = {
            "Все": None,
            "Руководство": ["Руководство"],
            "Производство": ["Смена 1", "Смена 2", "Прокат", "Резка"],
            "Склад": ["Склад"]
        }
        group_filter = groups.get(selected_group)

        # Получение списка сотрудников
        all_employees = self.employee_manager.get_all_employees()
        employees = [
            emp for emp in all_employees
            if emp['status'] == 'активный' and (group_filter is None or emp['department'] in group_filter)
        ]

        # Загрузка данных табеля
        timesheet = self.employee_manager.get_timesheet(year, month)
        self.timesheet_data = {
            (entry['employee_id'], datetime.strptime(entry['date'], '%Y-%m-%d').day): entry
            for entry in timesheet
        }
        employees_with_records = []
        for employee in employees:
            if employee['status'] == 'активный':
                employees_with_records.append(employee)
            elif employee['status'] == 'уволен':
                if self.employee_manager.has_timesheet_records(employee['id'], current_year_month):
                    employees_with_records.append(employee)

        self.table.clear()
        self.table.setRowCount(len(employees_with_records))  # Для каждого сотрудника одна строка
        self.table.setColumnCount(days_in_month + 1)  # Количество дней в месяце + 1 для суммы отработанных часов

        # Устанавливаем заголовки для дней месяца
        self.table.setHorizontalHeaderLabels(
            [str(day) for day in range(1, days_in_month + 1)] + ["Сумма часов"]
        )

        # Устанавливаем заголовки строк с фамилией и именем сотрудников
        self.table.setVerticalHeaderLabels(
            [f"{employee['lastname']} {employee['firstname']}" for employee in employees_with_records]
        )
        
        # Проверим, что timesheet_data формируется корректно
        self.timesheet_data = {
            (entry['employee_id'], datetime.strptime(entry['date'], '%Y-%m-%d').day): entry
            for entry in timesheet
        }

        for row, employee in enumerate(employees_with_records):
            total_hours = 0  # Переменная для накопления общей суммы отработанных часов
            shift_type = employee.get('department')

            for day in range(1, days_in_month + 1):  # Перебираем только дни, которые есть в текущем месяце
                self.table.setColumnWidth(day - 1, 75)  # Устанавливаем ширину для каждого дня
                date_key = (employee['id'], day)
                leave_status = self.get_leave_status(employee['id'], year, month, day)
                # Определяем текущую дату
                current_date = QDate(year, month, day)
                weekday = current_date.dayOfWeek()
                start_date = QDate(2024, 12, 30)
                # Логика для сменных графиков
                if shift_type == "Смена 1":
                    # 6 блоков для смены 1: 2 рабочих, 2 выходных, 3 рабочих, 2 выходных, 2 рабочих, 3 выходных
                    shift_cycle = [(2, False), (2, True), (3, False), (2, True), (2, False), (3, True)]
                elif shift_type == "Смена 2":
                    # 6 блоков для смены 2: инверсивно
                    shift_cycle = [(2, True), (2, False), (3, True), (2, False), (2, True), (3, False)]
                else:
                    shift_cycle = None

                if shift_cycle:
                    # Определяем день текущего месяца
                    current_date = QDate(year, month, day)

                    # Вычисляем разницу в днях между стартовой датой (30 декабря) и текущей датой
                    day_diff = start_date.daysTo(current_date)

                    # Вычисляем номер блока в цикле
                    cycle_length = sum([x[0] for x in shift_cycle])  # Длина всего цикла (сумма рабочих и выходных дней в цикле)
                    block_offset = day_diff % cycle_length  # Оставшийся день в текущем цикле

                    # Определяем день в блоке
                    for block in shift_cycle:
                        block_length, is_off_day = block
                        if block_offset < block_length:
                            # Если день в пределах текущего блока
                            is_working_day = not is_off_day  # Выходной или рабочий
                            break
                        else:
                            # Переходим к следующему блоку
                            block_offset -= block_length
                else:
                    # Пятидневка и праздники
                    weekday = current_date.dayOfWeek()
                    is_working_day = weekday not in {6, 7} and current_date.toPyDate() not in self.holidays

                if leave_status:
                    self.table.setItem(row, day - 1, QTableWidgetItem(leave_status))
                    self.table.item(row, day - 1).setBackground(Qt.gray)

                elif date_key in self.timesheet_data:
                    entry = self.timesheet_data[date_key]
                    arrival_time = entry['arrival_time']
                    departure_time = entry['departure_time']
                    
                    # Если время прихода существует, конвертируем его в строку времени
                    if arrival_time:
                        arrival_time_str = arrival_time
                    else:
                        arrival_time_str = "Нет данных"
                    
                    # Если время ухода существует, конвертируем его в строку времени
                    if departure_time:
                        departure_time_str = departure_time
                    else:
                        departure_time_str = "Нет данных"
                    
                    # Проверяем, что оба времени заданы
                    if arrival_time_str != "Нет данных" and departure_time_str != "Нет данных":
                        arrival_time_obj = QTime.fromString(arrival_time_str, "HH:mm")
                        departure_time_obj = QTime.fromString(departure_time_str, "HH:mm")
                        
                        if arrival_time_obj.isValid() and departure_time_obj.isValid():
                            worked_minutes = arrival_time_obj.msecsTo(departure_time_obj) / 60000
                            if worked_minutes > 240:
                                worked_minutes -= 30  # Учитываем обед
                            # Округляем время только при экспорте
                            rounded_minutes = round(worked_minutes / 30) * 30
                            worked_hours = rounded_minutes /60
                            total_hours += worked_hours

                            self.table.setItem(row, day - 1, QTableWidgetItem(f"{arrival_time_str} - {departure_time_str}"))
                            if not is_working_day:
                                self.table.item(row, day - 1).setBackground(QColor(255, 192, 203))
                        else:
                            self.table.setItem(row, day - 1, QTableWidgetItem("Некорректное время"))
                            if not is_working_day:
                                self.table.item(row, day - 1).setBackground(QColor(255, 192, 203))
                    else:
                        self.table.setItem(row, day - 1, QTableWidgetItem(f"{arrival_time_str} - "))
                        if not is_working_day:
                            self.table.item(row, day - 1).setBackground(QColor(255, 192, 203))
                else:
                    if not is_working_day:
                        self.table.setItem(row, day - 1, QTableWidgetItem("Выходной"))
                        self.table.item(row, day - 1).setBackground(QColor(255, 192, 203))
                    else:
                        self.table.setItem(row, day - 1, QTableWidgetItem("Нет данных"))

            # Записываем сумму отработанных часов в последнюю колонку
            self.table.setItem(row, days_in_month, QTableWidgetItem(f"{total_hours:.2f}"))

    def add_timesheet_entry(self):
        selected_date = self.calendar.selectedDate()
        dialog = TimesheetEntryDialog(self.employee_manager, selected_date, self)
        if dialog.exec_():
            self.load_timesheet()
    
    def export_to_excel(self):
        # Экспорт текущего табеля с учётом выбранной группы
        selected_group = self.group_selector.currentText()
        file_suffix = {"Все": "Все", "Руководство": "Руководство", "Производство": "Производство", "Склад": "Склад"}.get(selected_group, "Все")
        selected_date = self.calendar.selectedDate()
        year = selected_date.year()
        month = selected_date.month()
        current_year_month = f"{year}-{str(month).zfill(2)}"

        groups = {
            "Все": None,
            "Руководство": ["Руководство"],
            "Производство": ["Смена 1", "Смена 2", "Прокат", "Резка"],
            "Склад": ["Склад"]
        }
        group_filter = groups.get(selected_group)

        # Получаем всех сотрудников (активных и уволенных с записями)
        employees = self.employee_manager.get_all_employees()
        employees_with_records = [
        employee for employee in employees
        if (group_filter is None or employee['department'] in group_filter)
           and (employee['status'] == 'активный' or self.employee_manager.has_timesheet_records(employee['id'], current_year_month))
        ]

        # Обновляем self.timesheet_data только для сотрудников выбранной группы
        timesheet = self.employee_manager.get_timesheet(year, month)
        self.timesheet_data = {
            (entry['employee_id'], datetime.strptime(entry['date'], '%Y-%m-%d').day): entry
            for entry in timesheet if entry['employee_id'] in {emp['id'] for emp in employees_with_records}
        }

        # Создаем папку с названием года, если она не существует
        directory_name = os.path.join(r"C:\TimeTrackingSystem", str(year))
        if not os.path.exists(directory_name):
            os.makedirs(directory_name)

        # Путь к файлу Excel
        file_path = os.path.join(directory_name, f"{file_suffix}_{year}.xlsx")

        # Проверяем, существует ли файл
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
        else:
            workbook = Workbook()

        # Проверяем, существует ли лист с именем месяца
        month_name = selected_date.toString("MMMM")
        if month_name in workbook.sheetnames:
            sheet = workbook[month_name]
            # Очищаем данные на листе
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None
        else:
            sheet = workbook.create_sheet(title=month_name)

        # Устанавливаем заголовки столбцов
        sheet.cell(row=1, column=1, value="Сотрудник")
        days_in_month = selected_date.daysInMonth()
        for day in range(1, days_in_month + 1):
            sheet.cell(row=1, column=day + 1, value=str(day))
        sheet.cell(row=1, column=days_in_month + 2, value="Сумма часов")
        sheet.cell(row=1, column=days_in_month + 3, value="Заработная плата")

        # Заполняем данными таблицу
        for row, employee in enumerate(employees_with_records, start=2):
            sheet.cell(row=row, column=1, value=f"{employee['lastname']} {employee['firstname']}")
            total_hours = 0
            for day in range(1, days_in_month + 1):
                date_key = (employee['id'], day)
                leave_status = self.get_leave_status(employee['id'], year, month, day)
                if leave_status:
                    sheet.cell(row=row, column=day + 1, value=leave_status)
                    continue  # Пропускаем расчет времени, если статус — отпуск/больничный
                if date_key in self.timesheet_data:
                    entry = self.timesheet_data[date_key]
                    arrival_time = entry['arrival_time']
                    departure_time = entry['departure_time']
                    if arrival_time and departure_time:
                        arrival = QTime.fromString(arrival_time, "HH:mm")
                        departure = QTime.fromString(departure_time, "HH:mm")
                        if arrival.isValid() and departure.isValid():
                            worked_minutes = arrival.msecsTo(departure) / 60000
                            if worked_minutes > 240:
                                worked_minutes -= 30  # Учитываем обед
                            # Округляем время только при экспорте
                            rounded_minutes = round(worked_minutes / 30) * 30
                            worked_hours = rounded_minutes /60
                            total_hours += worked_hours
                            sheet.cell(row=row, column=day + 1, value=worked_hours)
                        else:
                            sheet.cell(row=row, column=day + 1, value="Некорр.")
                    else:
                        sheet.cell(row=row, column=day + 1, value="Нет данных")
            sheet.cell(row=row, column=days_in_month + 2, value=total_hours)
            salary = self.employee_manager.calculate_salary(employee['id'], year, month)
            sheet.cell(row=row, column=days_in_month + 3, value=salary)

        # Сохраняем файл
        workbook.save(file_path)
        workbook.close()

        QMessageBox.information(self, "Экспорт завершен", f"Данные успешно сохранены в файл:\n{file_path}")
 
    def get_leave_status(self, employee_id, year, month, day):
        """Проверяет, находится ли сотрудник на больничном или в отпуске в заданную дату."""
        leaves = self.employee_manager.get_leaves_for_employee(employee_id, year)
        date = QDate(year, month, day)
        for leave in leaves:
            start_date = QDate.fromString(leave[1], "yyyy-MM-dd")
            end_date = QDate.fromString(leave[2], "yyyy-MM-dd")
            if start_date <= date <= end_date:
                return "Больничный" if leave[0] == "Больничный" else "Отпуск"
        return None

class TimesheetEntryDialog(QDialog):
    def __init__(self, employee_manager, date, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Ввод данных табеля")
        self.setGeometry(200, 200, 300, 500)
        self.employee_manager = employee_manager
        self.date = date

        self.employee_selector = QListWidget()
        employees = self.employee_manager.get_all_employees()
        for employee in employees:
            item = QListWidgetItem(f"{employee['lastname']} {employee['firstname']}")
            item.setData(Qt.UserRole, employee['id'])
            self.employee_selector.addItem(item)

        self.arrival_time_input = QLineEdit()
        self.departure_time_input = QLineEdit()

        self.save_button = QPushButton("Сохранить")
        self.save_button.clicked.connect(self.save_entry)

        self.delete_button = QPushButton("Удалить")
        self.delete_button.clicked.connect(self.delete_entry)

        layout = QVBoxLayout()
        layout.addWidget(QLabel("Выберите сотрудника:"))
        layout.addWidget(self.employee_selector)
        layout.addWidget(QLabel("Время прихода (чч:мм):"))
        layout.addWidget(self.arrival_time_input)
        layout.addWidget(QLabel("Время ухода (чч:мм):"))
        layout.addWidget(self.departure_time_input)
        layout.addWidget(self.save_button)
        layout.addWidget(self.delete_button)

        self.setLayout(layout)

    def validate_time(self, time_str):
        """Проверяет, соответствует ли строка времени формату HH:mm."""
        if QTime.fromString(time_str, "HH:mm").isValid():
            return True
        QMessageBox.warning(self, "Ошибка", f"Некорректный формат времени: {time_str}. Используйте формат HH:mm.")
        return False

    def save_entry(self):
        selected_item = self.employee_selector.currentItem()
        if not selected_item:
            QMessageBox.warning(self, "Ошибка", "Выберите сотрудника.")
            return

        employee_id = selected_item.data(Qt.UserRole)
        arrival_time = self.arrival_time_input.text()
        departure_time = self.departure_time_input.text()

        # Проверка на корректный формат времени
        if arrival_time and not self.validate_time(arrival_time):
            return
        if departure_time and not self.validate_time(departure_time):
            return

        # Получаем текущие данные из базы для проверки
        timesheet_entry = self.employee_manager.get_timesheet_entry(employee_id, self.date.toString("yyyy-MM-dd"))

        # Если заполнены оба поля
        if arrival_time and departure_time:
            self.record_time(employee_id, arrival_time, departure_time)
            return

        # Если заполнено только поле "Приход"
        if arrival_time and not departure_time:
            if timesheet_entry and timesheet_entry['departure_time']:
                self.record_time(employee_id, arrival_time, timesheet_entry['departure_time'])
            else:
                self.record_time(employee_id, arrival_time, None)
            return

        # Если заполнено только поле "Уход"
        if departure_time and not arrival_time:
            if timesheet_entry and timesheet_entry['arrival_time']:
                self.record_time(employee_id, timesheet_entry['arrival_time'], departure_time)
            else:
                QMessageBox.warning(self, "Ошибка", "Введите время прихода перед добавлением времени ухода.")
            return

        # Если оба поля пусты
        QMessageBox.warning(self, "Ошибка", "Заполните хотя бы одно поле времени.")

    def record_time(self, employee_id, arrival_time, departure_time):
        """Метод для записи времени прихода и ухода сотрудника."""
        self.employee_manager.update_timesheet(
            employee_id,
            self.date.toString("yyyy-MM-dd"),
            arrival_time,
            departure_time
        )
        QMessageBox.information(self, "Успех", "Данные успешно сохранены.")
        #self.accept()

    def delete_entry(self):
        """Метод для удаления времени прихода и ухода для выбранного сотрудника и дня."""
        selected_item = self.employee_selector.currentItem()
        if not selected_item:
            QMessageBox.warning(self, "Ошибка", "Выберите сотрудника для удаления данных.")
            return

        employee_id = selected_item.data(Qt.UserRole)

        # Подтверждение удаления
        reply = QMessageBox.question(
            self, "Подтверждение удаления",
            "Вы уверены, что хотите удалить данные для выбранного сотрудника?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            self.employee_manager.remove_timesheet(employee_id, self.date.toString("yyyy-MM-dd"))
            QMessageBox.information(self, "Успех", "Данные успешно удалены.")
            self.accept()


class AddEmployeeDialog(QDialog):
    def __init__(self, employee_manager, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Добавление сотрудника")
        self.setGeometry(100, 100, 400, 300)
        self.employee_manager = employee_manager

        self.name_label = QLabel("Фамилия:")
        self.name_input = QLineEdit()

        self.firstname_label = QLabel("Имя:")
        self.firstname_input = QLineEdit()

        self.patronymic_label = QLabel("Отчество:")
        self.patronymic_input = QLineEdit()

        self.dob_label = QLabel("Дата рождения:")
        self.dob_input = QDateEdit()
        self.dob_input.setCalendarPopup(True)

        self.hire_date_label = QLabel("Дата приема на работу:")
        self.hire_date_input = QDateEdit()
        self.hire_date_input.setCalendarPopup(True)

        self.position_label = QLabel("Должность:")
        self.position_input = QLineEdit()

        self.save_button = QPushButton("Сохранить")
        self.save_button.clicked.connect(self.save_employee)

        self.cancel_button = QPushButton("Отмена")
        self.cancel_button.clicked.connect(self.reject)

        layout = QVBoxLayout()
        layout.addWidget(self.name_label)
        layout.addWidget(self.name_input)
        layout.addWidget(self.firstname_label)
        layout.addWidget(self.firstname_input)
        layout.addWidget(self.patronymic_label)
        layout.addWidget(self.patronymic_input)
        layout.addWidget(self.dob_label)
        layout.addWidget(self.dob_input)
        layout.addWidget(self.hire_date_label)
        layout.addWidget(self.hire_date_input)
        layout.addWidget(self.position_label)
        layout.addWidget(self.position_input)
        layout.addWidget(self.save_button)
        layout.addWidget(self.cancel_button)

        self.setLayout(layout)

    def save_employee(self):
        lastname = self.name_input.text()
        firstname = self.firstname_input.text()
        patronymic = self.patronymic_input.text()
        dob = self.dob_input.date().toString("yyyy-MM-dd")
        hire_date = self.hire_date_input.date().toString("yyyy-MM-dd")
        position = self.position_input.text()

        if not all([lastname, firstname, patronymic, dob, hire_date, position]):
            QMessageBox.warning(self, "Недостаточно данных", "Заполните все поля.")
            return

        self.employee_manager.add_employee(lastname, firstname, patronymic, dob, hire_date, position)
        
        # Закрываем диалоговое окно
        self.accept()

class EditEmployeeDialog(QDialog):
    def __init__(self, employee_manager, employee_id, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Редактирование сотрудника")
        self.setGeometry(100, 100, 400, 300)
        self.employee_manager = employee_manager
        self.employee_id = employee_id
        
        # Получаем данные сотрудника
        employee = self.employee_manager.get_employee_by_id(employee_id)
        leave_days = self.employee_manager.get_leave_days_for_year(employee_id, QDate.currentDate().year())

        self.name_label = QLabel("Фамилия:")
        self.name_input = QLineEdit(employee['lastname'])

        self.firstname_label = QLabel("Имя:")
        self.firstname_input = QLineEdit(employee['firstname'])

        self.patronymic_label = QLabel("Отчество:")
        self.patronymic_input = QLineEdit(employee['patronymic'])

        self.dob_label = QLabel("Дата рождения:")
        self.dob_input = QDateEdit()
        dob_str = employee['dob']
        dob_date = datetime.strptime(dob_str, "%Y-%m-%d")
        self.dob_input.setDate(QDate(dob_date.year, dob_date.month, dob_date.day))
        self.dob_input.setCalendarPopup(True)

        self.hire_date_label = QLabel("Дата приема на работу:")
        self.hire_date_input = QDateEdit()
        hire_date_str = employee['hire_date']
        hire_date_obj = datetime.strptime(hire_date_str, "%Y-%m-%d")
        self.hire_date_input.setDate(QDate(hire_date_obj.year, hire_date_obj.month, hire_date_obj.day))
        self.hire_date_input.setCalendarPopup(True)

        self.status_checkbox = QCheckBox("Уволен")
        self.status_checkbox.setChecked(employee['status'] == 'уволен')

        self.department_label = QLabel("Подразделение:")
        self.department_input = QComboBox()
        self.department_input.addItems(["Руководство", "Смена 1", "Смена 2", "Прокат", "Резка", "Склад"])
        self.department_input.setCurrentText(employee['department'])

        self.position_label = QLabel("Должность:")
        self.position_input = QLineEdit(employee['position'])

        self.bid_label = QLabel("Ставка:")
        self.bid_input = QLineEdit(str(employee['wages']))

        # Отображение дней в отпуске и на больничном
        self.vacation_days_label = QLabel(f"Дней в отпуске за текущий год: {leave_days['Отпуск']}")
        self.sick_days_label = QLabel(f"Дней на больничном за текущий год: {leave_days['Больничный']}")

        self.photo_label = QLabel("Фото:")
        self.photo_input = QLabel()
        self.photo_input.setGeometry(10, 10, 150, 150)  # Пример размеров QLabel
        self.photo_button = QPushButton("Выбрать фото")
        self.photo_button.clicked.connect(self.select_photo)

        self.AddFingerprint_button = QPushButton("Зарегестрировать отпечатки")
        self.AddFingerprint_button.clicked.connect(self.AddFingerprint)

        self.add_leave_button = QPushButton("Добавить отпуск/больничный")
        self.add_leave_button.clicked.connect(self.open_add_leave_window)

        self.save_button = QPushButton("Сохранить")
        self.save_button.clicked.connect(self.save_employee)

        self.delete_button = QPushButton("Удалить сотрудника")
        self.delete_button.clicked.connect(self.delete_employee)

        self.cancel_button = QPushButton("Отмена")
        self.cancel_button.clicked.connect(self.reject)

        layout = QVBoxLayout()
        layout.addWidget(self.name_label)
        layout.addWidget(self.name_input)
        layout.addWidget(self.firstname_label)
        layout.addWidget(self.firstname_input)
        layout.addWidget(self.patronymic_label)
        layout.addWidget(self.patronymic_input)
        layout.addWidget(self.dob_label)
        layout.addWidget(self.dob_input)
        layout.addWidget(self.hire_date_label)
        layout.addWidget(self.hire_date_input)
        layout.addWidget(self.department_label)
        layout.addWidget(self.department_input)
        layout.addWidget(self.position_label)
        layout.addWidget(self.position_input)
        layout.addWidget(self.bid_label)
        layout.addWidget(self.bid_input)
        layout.addWidget(self.status_checkbox)
        layout.addWidget(self.vacation_days_label)
        layout.addWidget(self.sick_days_label)
        layout.addWidget(self.photo_label)
        layout.addWidget(self.photo_input)
        layout.addWidget(self.photo_button)
        layout.addWidget(self.AddFingerprint_button)
        layout.addWidget(self.add_leave_button)
        layout.addWidget(self.save_button)
        layout.addWidget(self.delete_button)
        layout.addWidget(self.cancel_button)

        self.setLayout(layout)
        
        # Загружаем фото сотрудника, если оно есть
        self.photo_input = self.load_employee_photo(employee['photo'])

    def get_days_in_leave(self, leave_type):
        current_year = QDate.currentDate().year()
        leaves = self.employee_manager.get_leaves_for_employee(self.employee_id, current_year)
        total_days = 0
        for leave in leaves:
            if leave[0] == leave_type:
                start_date = QDate.fromString(leave[1], "yyyy-MM-dd")
                end_date = QDate.fromString(leave[2], "yyyy-MM-dd")
                total_days += start_date.daysTo(end_date) + 1
        return total_days

    def open_add_leave_window(self):
        dialog = AddLeaveDialog(self.employee_manager, self.employee_id, self)
        dialog.exec_()

    def load_employee_photo(self, photo_data):
        if photo_data:
            # Отображаем фото в photo_input
            self.display_photo(photo_data, self.photo_input)

    def display_photo(self, photo_data, input):
        """
        Отображение фото из бинарных данных в QLabel.
        :param photo_data: Бинарные данные фото из базы данных.
        :param input: QLabel для отображения изображения.
        """
        if photo_data:
            pixmap = QPixmap()
            if pixmap.loadFromData(photo_data):
                input.setPixmap(pixmap.scaled(150, 150, Qt.KeepAspectRatio, Qt.SmoothTransformation))
                input.setAlignment(Qt.AlignCenter)  # Выравнивание изображения по центру QLabel
            else:
                logging.error("Не удалось загрузить фото из данных.")
                input.clear()  # Очистить, если загрузка не удалась
        else:
            input.clear()  # Очистить, если данных нет

    def select_photo(self):
        # Открываем диалог для выбора фото
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "Выбрать фото", "", "Image Files (*.png *.jpg *.bmp)", options=options)
        if file_name:
            self.photo_path = file_name
            # Здесь можно добавить обработку обрезки фото
            self.edit_photo()

    def edit_photo(self):
        # Открываем окно редактирования фото (обрезка и масштабирование)
        self.photo_editor = PhotoEditor(self.photo_path, self.employee_id)
        self.photo_editor.exec_()
    
    def AddFingerprint(self):
        Fingerprint = AddFingerprintDialog(self.employee_id, self.employee_manager)
        Fingerprint.exec_()
    
    def save_employee(self):
        lastname = self.name_input.text().strip()
        firstname = self.firstname_input.text().strip()
        patronymic = self.patronymic_input.text().strip()
        dob = self.dob_input.date().toString("yyyy-MM-dd")
        hire_date = self.hire_date_input.date().toString("yyyy-MM-dd")
        position = self.position_input.text().strip()
        wages = self.bid_input.text().strip()
        department = self.department_input.currentText().strip()

        # Проверяем, что все обязательные поля заполнены
        if not all([lastname, firstname, patronymic, dob, hire_date, position, wages]):
            QMessageBox.warning(self, "Недостаточно данных", "Заполните все поля.")
            return

        # Сохраняем изменения в базе данных
        status = 'уволен' if self.status_checkbox.isChecked() else 'активный'
        self.employee_manager.update_employee(self.employee_id, lastname, firstname, patronymic, dob, hire_date, position, wages, status, department)

        logging.info(f"Данные сотрудника {lastname} {firstname} сохранены.")
        self.accept()

    def delete_employee(self):
        self.employee_manager.delete_employee(self.employee_id)
        QMessageBox.information(self, "Успех", "Сотрудник удален.")
        self.accept()

class AddLeaveDialog(QDialog):
    def __init__(self, employee_manager, employee_id, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Добавить отпуск/больничный")
        self.setGeometry(200, 200, 300, 200)
        self.employee_manager = employee_manager
        self.employee_id = employee_id

        self.type_label = QLabel("Тип:")
        self.type_selector = QComboBox()
        self.type_selector.addItems(["Отпуск", "Больничный"])

        self.start_date_label = QLabel("Дата начала:")
        self.start_date_input = QDateEdit()
        self.start_date_input.setCalendarPopup(True)
        self.start_date_input.setDate(QDate.currentDate())

        self.end_date_label = QLabel("Дата окончания:")
        self.end_date_input = QDateEdit()
        self.end_date_input.setCalendarPopup(True)
        self.end_date_input.setDate(QDate.currentDate())

        self.save_button = QPushButton("Сохранить")
        self.save_button.clicked.connect(self.save_leave)

        self.cancel_button = QPushButton("Отмена")
        self.cancel_button.clicked.connect(self.reject)

        layout = QVBoxLayout()
        layout.addWidget(self.type_label)
        layout.addWidget(self.type_selector)
        layout.addWidget(self.start_date_label)
        layout.addWidget(self.start_date_input)
        layout.addWidget(self.end_date_label)
        layout.addWidget(self.end_date_input)
        layout.addWidget(self.save_button)
        layout.addWidget(self.cancel_button)

        self.setLayout(layout)

    def save_leave(self):
        leave_type = self.type_selector.currentText()
        start_date = self.start_date_input.date().toString("yyyy-MM-dd")
        end_date = self.end_date_input.date().toString("yyyy-MM-dd")

        if start_date > end_date:
            QMessageBox.warning(self, "Ошибка", "Дата начала не может быть позже даты окончания.")
            return

        self.employee_manager.add_leave(self.employee_id, start_date, end_date, leave_type)
        QMessageBox.information(self, "Успех", "Запись об отпуске/больничном добавлена.")
        self.accept()

class PhotoEditor(QDialog):
    def __init__(self, image_path, employee_id, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Редактор фото сотрудника")
        self.setGeometry(100, 100, 800, 600)

        self.image_path = image_path
        self.employee_id = employee_id
        self.original_pixmap = QPixmap(image_path)
        self.current_pixmap = self.original_pixmap.copy()

        # Текущие параметры
        self.rotation_angle = 0  # Текущий угол поворота
        self.zoom_factor = 1.0  # Коэффициент масштабирования

        # Создаем интерфейс
        self.init_ui()

    def init_ui(self):
        """Инициализация пользовательского интерфейса"""
        self.graphics_view = QGraphicsView(self)
        self.scene = QGraphicsScene(self)
        self.graphics_view.setScene(self.scene)
        
        # Создаем изображение и добавляем в сцену
        self.image_item = QGraphicsPixmapItem(self.current_pixmap)
        self.scene.addItem(self.image_item)

        # Прямоугольник для обрезки
        self.crop_rect = QGraphicsRectItem(0, 0, 200, 200)
        self.crop_rect.setFlag(QGraphicsRectItem.ItemIsMovable)
        self.crop_rect.setFlag(QGraphicsRectItem.ItemIsSelectable)
        self.crop_rect.setPen(Qt.red)
        self.scene.addItem(self.crop_rect)

        # Слайдер для изменения масштаба
        self.zoom_slider = QSlider(Qt.Horizontal)
        self.zoom_slider.setMinimum(10)
        self.zoom_slider.setMaximum(200)
        self.zoom_slider.setValue(100)
        self.zoom_slider.valueChanged.connect(self.update_zoom)

        # Кнопки управления
        self.rotate_button = QPushButton("Повернуть", self)
        self.rotate_button.clicked.connect(self.rotate_image)

        self.save_button = QPushButton("Сохранить", self)
        self.save_button.clicked.connect(self.save_cropped_image)

        self.cancel_button = QPushButton("Отмена", self)
        self.cancel_button.clicked.connect(self.reject)

        # Макет кнопок
        button_layout = QHBoxLayout()
        button_layout.addWidget(self.rotate_button)
        button_layout.addWidget(self.save_button)
        button_layout.addWidget(self.cancel_button)

        layout = QVBoxLayout()
        layout.addWidget(self.graphics_view)
        layout.addWidget(self.zoom_slider)
        layout.addLayout(button_layout)
        self.setLayout(layout)

    def apply_transformations(self):
        """Применение текущих трансформаций (поворот и масштаб)"""
        transform = QTransform()
        transform.rotate(self.rotation_angle)
        scaled_pixmap = self.original_pixmap.transformed(transform, Qt.SmoothTransformation)
        
        # Применяем масштабирование
        width = int(scaled_pixmap.width() * self.zoom_factor)
        height = int(scaled_pixmap.height() * self.zoom_factor)
        scaled_pixmap = scaled_pixmap.scaled(
            width,
            height,
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation
        )
        self.current_pixmap = scaled_pixmap  # Сохраняем изменённое изображение в переменную
        self.image_item.setPixmap(self.current_pixmap)  # Обновляем отображаемое изображение

    def update_zoom(self):
        """Обновить масштаб изображения в зависимости от значения слайдера"""
        self.zoom_factor = self.zoom_slider.value() / 100.0
        self.apply_transformations()

    def rotate_image(self):
        """Повернуть изображение на 90 градусов"""
        self.rotation_angle = (self.rotation_angle + 90) % 360
        self.apply_transformations()

    def save_cropped_image(self):
        """Сохранить обрезанное изображение с учетом смещения рамки"""
        # Получаем прямоугольник рамки относительно сцены
        scene_rect = self.crop_rect.sceneBoundingRect()

        # Получаем прямоугольник изображения относительно сцены
        image_rect = self.image_item.sceneBoundingRect()

        # Преобразуем координаты рамки в координаты изображения
        rect = scene_rect.translated(-image_rect.topLeft())  # Сдвигаем координаты рамки относительно изображения

        # Обрезаем изображение по корректному прямоугольнику
        cropped_image = self.image_item.pixmap().copy(rect.toRect())

        # Применяем трансформации (поворот, масштабирование)
        self.apply_transformations()

        # Сохраняем изображение в файл для проверки
        #self.save_image_to_file(cropped_image)
        
        # Преобразуем изображение в байтовый массив
        byte_array = self.image_to_byte_array(cropped_image)
        
        # Сохраняем изображение в базе данных
        self.save_image_to_db(byte_array)

        # Закрываем окно
        self.accept()

    #def save_image_to_file(self, image: QPixmap):
    #    """Сохранение изображения в файл для проверки"""
    #    file_path = "cropped_image.png"  # Указываем путь, по которому сохраняем
    #    image.save(file_path, "PNG")  # Сохраняем изображение в формате PNG
    #    logging.info(f"Изображение сохранено в файл: {file_path}")

    def image_to_byte_array(self, image: QPixmap):
        """Преобразование QPixmap в байтовый массив"""
        byte_array = QByteArray()
        buffer = QBuffer(byte_array)
        buffer.open(QIODevice.WriteOnly)
        image.save(buffer, format="PNG")
        return byte_array.data()

    def save_image_to_db(self, image_data):
        """Сохранение изображения в базе данных SQLite"""
        try:
            # Установка соединения с базой данных SQLite
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()

            # Проверка содержимого перед вставкой
            if not image_data:
                logging.error("Ошибка: нет данных изображения.")
                return

            # SQL-запрос для обновления фото сотрудника
            query = "UPDATE employees SET photo = ? WHERE id = ?"
            cursor.execute(query, (image_data, self.employee_id))
            conn.commit()
            cursor.close()
            conn.close()

        except sqlite3.Error as err:
            logging.error(f"Ошибка при сохранении фото в базе данных: {err}")

class AddFingerprintDialog(QDialog):
    def __init__(self, employee_id, employee_manager, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Выбор пальца для регистрации отпечатка")
        self.setFixedSize(450, 300)  # Установим фиксированный размер окна
        self.employee_id = employee_id
        self.employee_manager = employee_manager

        # Загружаем зарегистрированные пальцы
        self.registered_fingers = self.load_registered_fingers()

        # Основной лейаут
        main_layout = QVBoxLayout()

        # Верхняя надпись
        instructions = QLabel("Выберите палец для регистрации отпечатка:")
        instructions.setAlignment(Qt.AlignCenter)
        instructions.setGeometry(10, 10, 400, 20)
        instructions.setParent(self)

        # Установка фонового изображения
        self.background_label = QLabel(self)
        pixmap = QPixmap("media/images/hands2.png")  # Укажите путь к изображению
        pixmap = pixmap.scaled(450, 240, Qt.KeepAspectRatio)
        self.background_label.setPixmap(pixmap)
        self.background_label.setGeometry(0, 30, 450, 240)

        # Радиокнопки с их точными координатами на изображении
        radio_buttons = {
            "LF5": (25, 86), "LF4": (60, 55), "LF3": (108, 37), "LF2": (158, 62), "LF1": (182, 143),
            "RF1": (256, 143), "RF2": (281, 62), "RF3": (330, 37), "RF4": (378, 55), "RF5": (413, 86),
        }
        self.radio_buttons = {}

        # Добавление радиокнопок
        for finger_name, (x, y) in radio_buttons.items():
            self.add_radio_button(finger_name, x, y, self.registered_fingers)

        # Установка радиокнопки LF1 по умолчанию
        if "LF1" in self.radio_buttons:
            self.radio_buttons["LF1"].setChecked(True)

        # Кнопка подтверждения
        confirm_button = QPushButton("Подтвердить", self)
        confirm_button.setGeometry(125, 260, 100, 30)
        confirm_button.clicked.connect(self.confirm_selection)

        # Кнопка для закрытия окна
        close_button = QPushButton("Закрыть", self)
        close_button.setGeometry(235, 260, 100, 30)
        close_button.clicked.connect(self.reject)

        self.setLayout(main_layout)

    def add_radio_button(self, finger_name, x, y, registered_fingers):
        #Добавляет радиокнопку с информацией о качестве.
        radio_button = QRadioButton(self)
        radio_button.move(x, y)
        self.radio_buttons[finger_name] = radio_button

        # Проверяем, зарегистрирован ли палец, и добавляем информацию о качестве
        quality = registered_fingers.get(finger_name, None)
        label_text = f"{finger_name} ({quality}/10)" if quality else finger_name
        label = QLabel(label_text, self)
        label.move(x - 43, y - 15)
        label.setAlignment(Qt.AlignCenter)

    def confirm_selection(self):
        #Подтверждение выбора пальца.
        selected_finger = None
        for finger_name, radio_button in self.radio_buttons.items():
            if radio_button.isChecked():
                selected_finger = finger_name
                break

        if selected_finger:
            self.start_scan(selected_finger)

    def load_registered_fingers(self):
        #Загрузка информации о зарегистрированных пальцах из базы данных SQLite.
        try:
            # Открываем соединение с базой данных SQLite
            cursor = self.employee_manager.db_connection.cursor()

            # Выполняем запрос для получения данных о зарегистрированных пальцах
            cursor.execute("""
                SELECT finger_name, quality 
                FROM fingerprints 
                WHERE employee_id = ?
            """, (self.employee_id,))
            
            # Получаем результаты и формируем словарь
            results = cursor.fetchall()
            return {row[0]: row[1] for row in results}

        except sqlite3.Error as err:
            logging.error(f"Ошибка при загрузке зарегистрированных пальцев: {err}")
            return {}

    def start_scan(self, finger_name):
        #Открывает окно для сканирования выбранного пальца.
        self.accept()  # Закрываем окно выбора
        self.scan_finger_dialog = FingerprintRegistrationDialog(self.employee_manager, self.employee_id, finger_name, self)
        self.scan_finger_dialog.exec_()

class FingerprintRegistrationDialog(QDialog):
    registration_complete = pyqtSignal(int, str, bytes, int)  # employee_id, finger_name, template, quality
    def __init__(self, employee_manager, employee_id, finger_name, parent=None):
        global cancel_response_flag
        cancel_response_flag = False
        super().__init__(parent)
        self.sdk = FutronicSDK()
        self.setWindowTitle("Регистрация отпечатка")
        self.setGeometry(100, 100, 400, 300)
        self.employee_id = employee_id
        self.finger_name = finger_name
        self.employee_manager = employee_manager
        # Элементы для отображения отпечатка и статуса
        self.fingerprint_image = QLabel()  # Здесь будет отображаться изображение отпечатка

        self.status_text = QTextEdit()
        self.status_text.setReadOnly(True)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)

        self.stop_button = QPushButton("Выбрать другой палец")
        self.stop_button.clicked.connect(self.stop)

        self.close_button = QPushButton("Закрыть")
        self.close_button.clicked.connect(self.close)

        layout = QVBoxLayout()

        layout1 = QHBoxLayout()
        layout1.addWidget(self.fingerprint_image)
        layout1.addWidget(self.status_text)
        layout.addLayout(layout1)
        layout.addWidget(self.progress_bar)
        layout.addWidget(self.stop_button)
        layout.addWidget(self.close_button)

        self.setLayout(layout)
        self.sdk.registration_complete.connect(self.on_registration_complete)

        # Создаем и запускаем поток сканирования
        self.scanning_thread = ScanningThread(self.sdk, employee_id, finger_name,)
        self.scanning_thread.update_status.connect(self.update_status)
        self.scanning_thread.update_image.connect(self.update_fingerprint_image)
        self.scanning_thread.registration_complete.connect(self.on_registration_complete)
        #Подключаем сигнал SDK к статусному текстовому полю
        self.sdk.status_signal.connect(self.update_status)
        self.sdk.status_image.connect(self.update_fingerprint_image)
        self.sdk.finger_lift_signal.connect(self.update_progress_bar)
        # Начинаем процесс сканирования
        self.scanning_thread.start()

    def stop(self):
        global cancel_response_flag
        cancel_response_flag = True
        QTimer.singleShot(1000, self.after_delay_stop)

    def after_delay_stop(self):
        self.accept()
        Fingerprint = AddFingerprintDialog(self.employee_id, self.employee_manager)
        Fingerprint.exec_()
    
    # Переопределяем обработчик закрытия окна
    def closeEvent(self, event):
        global cancel_response_flag
        cancel_response_flag = True
        QTimer.singleShot(1000, self.after_delay)
        
    def after_delay(self, event):
        event.accept()  # Разрешить закрытие окна

    def update_status(self, message):
        if self.sdk:  # Убедитесь, что SDK не завершен
            self.status_text.append(message)

    def update_fingerprint_image(self, pixmap):
        pixmap = pixmap.scaled(133, 175, Qt.KeepAspectRatio)
        self.fingerprint_image.setPixmap(pixmap)

    def update_progress_bar(self):
        current_value = self.progress_bar.value()
        increment = 34  # 33% для каждого сообщения
        self.progress_bar.setValue(min(current_value + increment, 100))  # Не больше 100%

    def handle_registration_complete(self):
        self.status_text.append("Регистрация завершена. Вы можете закрыть окно.")

    def handle_thread_finished(self):
        self.status_text.append("Процесс завершён.")

    def on_registration_complete(self, employee_id, finger_name, fingerprint_template, quality):
        #Обработчик завершения регистрации отпечатка.
        self.employee_manager.add_or_update_fingerprint(employee_id, finger_name, fingerprint_template, quality)
        self.status_text.append(f"Отпечаток для {finger_name} сохранен с качеством {quality}/10.")
        self.accept()



class ScanningThread(QThread):
    update_status = pyqtSignal(int)
    update_image = pyqtSignal(str)
    registration_complete = pyqtSignal(int, str, bytes, int)

    def __init__(self, sdk, employee_id, finger_name, parent=None):
        super().__init__(parent)
        self.sdk = sdk
        self.running = True
        self.employee_id = employee_id
        self.finger_name = finger_name
    def run(self):
        try:
            if self.running:
                self.sdk.Initialize()
                self.sdk.Enrollment(self.employee_id, self.finger_name)
        except Exception as e:
            logging.error(f"Ошибка: {e}")
        finally:
            self.sdk.Terminate()
            self.running = False

    def stop(self):
        self.running = False
        self.sdk.Terminate()
        self.wait()
        self.quit()

class FtrData(Structure):
        _pack_ = 1
        _fields_ = [('dwsize', c_ulong),
                    ('pdata', POINTER(c_void_p))]

class FtrBitmap(Structure):
        _pack_ = 1
        _fields_ = [('width', c_ulong),
                    ('height', c_ulong),
                    ('bitmap', FtrData)]

class FtrEnrollData(Structure):
        _pack_ = 1
        _fields_ = [('dwsize', c_ulong),
                    ('dwquality', c_ulong)]

class FtrIdentifyRecord(Structure):
        _pack_ = 1
        _fields_ = [('keyvalue', c_char * 16),
                    ('pdata', POINTER(FtrData))]

class FtrIdentifyArray(Structure):
        _pack_ = 1
        _fields_ = [('TotalNumber', c_ulong), ('pmembers', POINTER(FtrIdentifyRecord))]

class FarAttained(Union):
        _pack_ = 1
        _fields_ = [('p', c_ulong),
                    ('n', c_long)]

class FtrMatchedXRecord(Structure):
        _pack_ = 1
        _fields_ = [('keyvalue', c_char * 16),
                    ('far_attained', FarAttained)]

class FtrMatchedXArray(Structure):
        _pack_ = 1
        _fields_ = [('TotalNumber', c_ulong), ('pmembers', POINTER(FtrMatchedXRecord))]

# Глобальный флаг для управления отменой
cancel_response_flag = False
class FutronicSDK(QObject):
    update_status = pyqtSignal(int)
    status_signal = pyqtSignal(str)
    status_image = pyqtSignal(QPixmap)
    finger_lift_signal = pyqtSignal()
    registration_complete = pyqtSignal(int, str, bytes, int)
    ftrdll = CDLL(DLL_PATH, mode=RTLD_GLOBAL)

    def __init__(self):
        super().__init__()
        self.baseSample = FtrData()
        self.enrolSample = FtrData()
        self.cb_control_func = self.create_cb_control()  # Создаем обертку

    def create_cb_control(self):
        callback = CFUNCTYPE(None, 
            FTR_USER_CTX,  # Ввод: контекст пользователя
            FTR_STATE,     # Ввод: маска состояния
            POINTER(FTR_RESPONSE),  # Вывод: управление функцией API
            FTR_SIGNAL,    # Ввод: сигнал взаимодействия
            FTR_BITMAP_PTR # Ввод: указатель на bitmap
            )
        def control(context, state, response, signal, bitmap):
            global cancel_response_flag
            # Проверяем флаг для отмены
            if cancel_response_flag:
                response.contents.value = FTR_CB_RESP_CANCEL.value
                
            else:
                response.contents.value = FTR_CB_RESP_CONTINUE.value
                def signal_undefined():
                    self.status_signal.emit("Неопределенный сигнал...")

                def signal_touch():
                    self.status_signal.emit("Приложите палец к сканеру.")
                    self.status_image.emit(QPixmap("media/images/put your finger.png"))

                def signal_takeoff():
                    self.status_signal.emit("Уберите палец с сканера.")
                    self.status_image.emit(QPixmap("media/images/remove your finger.png"))
                    self.finger_lift_signal.emit()

                def signal_fakesource():
                    self.status_signal.emit("Обнаружен ложный источник сигнала.")

                switch = {
                    0: signal_undefined,
                    1: signal_touch,
                    2: signal_takeoff,
                    3: signal_fakesource
                }

                if state & FTR_STATE_SIGNAL_PROVIDED:
                    switch[signal]()
        callback_function = callback(control)
        return callback_function

    def Initialize(self):
        res = c_long(0)
        res = self.ftrdll.FTRInitialize()
        res = self.ftrdll.FTRSetParam(FTR_PARAM_CB_FRAME_SOURCE, FSD_FUTRONIC_USB)
        res = self.ftrdll.FTRSetParam(FTR_PARAM_CB_CONTROL, self.cb_control_func)
        res = self.ftrdll.FTRSetParam(FTR_PARAM_MAX_FARN_REQUESTED, 245)
        res = self.ftrdll.FTRSetParam(FTR_PARAM_FAKE_DETECT, False)
        res = self.ftrdll.FTRSetParam(FTR_PARAM_FFD_CONTROL, False)
        res = self.ftrdll.FTRSetParam(FTR_PARAM_MIOT_CONTROL, True)
        res = self.ftrdll.FTRSetParam(FTR_PARAM_MAX_MODELS, 3)
        res = self.ftrdll.FTRSetParam(FTR_PARAM_VERSION, FTR_VERSION_CURRENT)
        res = self.ftrdll.FTRGetParam(FTR_PARAM_MAX_TEMPLATE_SIZE, byref(self.enrolSample, FtrData.dwsize.offset))

    def Enrollment(self, employee_id, finger_name):
        self.employee_id = employee_id
        self.finger_name = finger_name
        #1. Enrollment: 
        self.status_signal.emit("Сканирование...")
        self.baseSample.pdata = (c_void_p * self.enrolSample.dwsize)()
        self.enrolSample.pdata = (c_void_p * self.enrolSample.dwsize)()
        eData = FtrEnrollData()
        eData.dwsize = sizeof(FtrEnrollData)
        result = self.ftrdll.FTREnrollX(None, FTR_PURPOSE_ENROLL, byref(self.enrolSample), byref(eData))

        if result == FTR_RETCODE_OK:
            self.status_signal.emit("Сканирование завершено.")

            # 3. Преобразование шаблона в массив байт
            array_type = c_ubyte * self.enrolSample.dwsize
            data_carray = array_type.from_address(addressof(self.enrolSample.pdata.contents))
            fingerprint_template = bytes(data_carray)
            # Сохранение отпечатка
            self.registration_complete.emit(self.employee_id, self.finger_name, fingerprint_template, eData.dwquality)
        elif result == 8:
            self.status_signal.emit("Сканирование отменено.")
        else:
            self.status_signal.emit("Ошибка при сканировании.")
        self.Terminate()

    def Terminate(self):
        self.ftrdll.FTRTerminate()



class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Система управления сотрудниками")
        self.setGeometry(100, 100, 800, 600)

        self.employee_manager = EmployeeManager()

        self.add_button = QPushButton("Добавить сотрудника")
        self.add_button.clicked.connect(self.add_employee)

        self.show_dismissed_button = QPushButton("Показать уволенных сотрудников")
        self.show_dismissed_button.clicked.connect(self.show_dismissed_employees)

        self.timesheet_button = QPushButton("Просмотр табеля")
        self.timesheet_button.clicked.connect(self.show_timesheet)

        self.employee_list = QListWidget()
        self.employee_list.itemClicked.connect(self.on_employee_clicked)

        self.load_employee_list()

        layout = QVBoxLayout()
        layout.addWidget(self.add_button)
        layout.addWidget(self.show_dismissed_button)
        layout.addWidget(self.timesheet_button)
        layout.addWidget(self.employee_list)

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)
        self.update_timer = QTimer(self)
        self.update_timer.timeout.connect(self.load_employee_list)
        self.update_timer.start(30000)  # Обновление каждые 30 секунд


    def show_dismissed_employees(self):
        dialog = DismissedEmployeesWindow(self.employee_manager, self)
        dialog.exec_()

    def load_employee_list(self):
        self.employee_list.clear()
        employees = [emp for emp in self.employee_manager.get_all_employees() if emp['status'] == 'активный']
        current_date = QDate.currentDate().toString("yyyy-MM-dd")
        for emp in employees:
            item_widget = QWidget()
            layout = QHBoxLayout()

            # Фото сотрудника
            photo_data = self.employee_manager.get_employee_by_id(emp['id']).get('photo')
            photo_label = QLabel()
            if photo_data:
                pixmap = QPixmap()
                if pixmap.loadFromData(photo_data):
                    photo_label.setPixmap(pixmap.scaled(50, 50, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            photo_label.setFixedSize(50, 50)

            # Текстовая информация
            employee_label = QLabel(f"{emp['lastname']} {emp['firstname']} {emp['patronymic']}")
            employee_label.setFixedSize(250, 50)

            # Проверка записей о приходе и уходе
            timesheet_entry = self.employee_manager.get_timesheet_entry(emp['id'], current_date)

            job_status =  QLabel()

            leaves = self.employee_manager.get_leaves_for_employee(emp['id'], QDate.currentDate().year())
            is_on_leave = False
            for leave in leaves:
                start_date = QDate.fromString(leave[1], "yyyy-MM-dd")
                end_date = QDate.fromString(leave[2], "yyyy-MM-dd")
                if start_date <= QDate.currentDate() <= end_date:
                    is_on_leave = True
                    if leave[0] == "Отпуск":
                        job_status.setText("В отпуске")
                        job_status.setStyleSheet("color: blue; font-weight: bold;")
                    elif leave[0] == "Больничный":
                        job_status.setText("На больничном")
                        job_status.setStyleSheet("color: orange; font-weight: bold;")
                    break

            # Если сотрудник не на больничном и не в отпуске, проверяем состояние по табелю
            if not is_on_leave:
                if timesheet_entry:
                    if timesheet_entry['arrival_time'] and not timesheet_entry['departure_time']:
                        job_status.setText("На работе")
                        job_status.setStyleSheet("color: green; font-weight: bold;")
                    elif timesheet_entry['arrival_time'] and timesheet_entry['departure_time']:
                        job_status.setText("Отсутствует")
                        job_status.setStyleSheet("color: red; font-weight: bold;")
                else:
                    job_status.setText("Отсутствует")
                    job_status.setStyleSheet("color: red; font-weight: bold;")
            job_status.setFixedSize(80, 50)


            wages =  QLabel("ЗП:")
            wages.setFixedSize(30, 50)
            wages_text = QLabel()
            wages_text.setFixedSize(100, 50)
            employee_id = emp['id']
            current_year = QDate.currentDate().year()
            current_month = QDate.currentDate().month()
            salary = self.employee_manager.calculate_salary(employee_id, current_year, current_month)
            wages_text.setText(f"{salary:.2f} руб.")

            # Рассчет времени до дня рождения
            if emp['dob']:
                # Преобразуем строку в объект datetime
                dob_date = datetime.strptime(emp['dob'], '%Y-%m-%d')
                
                # Создаём объект QDate из datetime
                birth_date = QDate(dob_date.year, dob_date.month, dob_date.day)
                
                # Получаем текущую дату
                today = QDate.currentDate()
                
                # Рассчитываем следующую дату рождения
                next_birthday = QDate(today.year(), birth_date.month(), birth_date.day())
                
                # Если день рождения уже прошёл в этом году, то берём следующий год
                if today > next_birthday:
                    next_birthday = QDate(today.year() + 1, birth_date.month(), birth_date.day())
                
                # Рассчитываем количество дней до следующего дня рождения
                days_left = today.daysTo(next_birthday)
                
                # Выводим информацию о количестве дней до дня рождения
                birthday_label = QLabel(f"До дня рождения: {days_left} дней")
                birthday_label.setFixedSize(150, 50)
            else:
                birthday_label = QLabel("Дата рождения неизвестна")

            layout.addWidget(photo_label)
            layout.addWidget(employee_label)
            layout.addWidget(job_status)
            layout.addWidget(wages)
            layout.addWidget(wages_text)
            layout.addWidget(birthday_label)
            layout.addStretch()
            item_widget.setLayout(layout)

            # Добавление в список
            list_item = QListWidgetItem()
            list_item.setSizeHint(item_widget.sizeHint())
            list_item.setData(Qt.UserRole, emp['id'])
            self.employee_list.addItem(list_item)
            self.employee_list.setItemWidget(list_item, item_widget)

    def add_employee(self):
        dialog = AddEmployeeDialog(self.employee_manager, self)
        dialog.exec_()
        self.load_employee_list()

    def on_employee_clicked(self, item):
        emp_id = item.data(Qt.UserRole)
        if emp_id is not None:
            dialog = EditEmployeeDialog(self.employee_manager, emp_id, self)
            dialog.exec_()
            self.load_employee_list()

    def show_timesheet(self):
        dialog = TimesheetWindow(self.employee_manager, self)
        dialog.exec_()

class DismissedEmployeesWindow(QDialog):
    def __init__(self, employee_manager, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Уволенные сотрудники")
        self.setGeometry(100, 100, 600, 400)
        self.employee_manager = employee_manager

        # Список сотрудников
        self.employee_list = QListWidget()
        self.employee_list.itemClicked.connect(self.on_employee_clicked)

        # Загрузка списка уволенных сотрудников
        self.load_dismissed_employees()

        # Кнопка закрытия
        self.close_button = QPushButton("Закрыть")
        self.close_button.clicked.connect(self.reject)

        # Макет
        layout = QVBoxLayout()
        layout.addWidget(self.employee_list)
        layout.addWidget(self.close_button)
        self.setLayout(layout)

    def load_dismissed_employees(self):
        """Загружает список уволенных сотрудников."""
        employees = self.employee_manager.get_dismissed_employees()
        self.employee_list.clear()
        for emp in employees:
            item = QListWidgetItem(f"{emp['lastname']} {emp['firstname']} {emp['patronymic']}")
            item.setData(Qt.UserRole, emp['id'])
            self.employee_list.addItem(item)

    def on_employee_clicked(self, item):
        """Открывает карточку сотрудника при клике."""
        emp_id = item.data(Qt.UserRole)
        if emp_id is not None:
            dialog = EditEmployeeDialog(self.employee_manager, emp_id, self)
            dialog.exec_()
            self.load_dismissed_employees()

class PasswordDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Вход")
        self.setGeometry(100, 100, 300, 150)

        self.password_label = QLabel("Введите пароль:")
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)
        
        self.ok_button = QPushButton("OK")
        self.ok_button.clicked.connect(self.check_password)
        
        self.cancel_button = QPushButton("Отмена")
        self.cancel_button.clicked.connect(self.reject)

        layout = QVBoxLayout()
        layout.addWidget(self.password_label)
        layout.addWidget(self.password_input)
        layout.addWidget(self.ok_button)
        layout.addWidget(self.cancel_button)

        self.setLayout(layout)

    def check_password(self):
        password = self.password_input.text()
        # Пример проверки пароля (замените "admin" на нужный пароль)
        if password == "Efimov5427720":
            self.accept()
        else:
            logging.info(f"Неверный пароль: {password}")
            QMessageBox.warning(self, "Ошибка", "Неверный пароль!")

def main():
    app = QApplication(sys.argv)
        # Окно ввода пароля
    password_dialog = PasswordDialog()
    if password_dialog.exec_() != QDialog.Accepted:
        sys.exit()  # Закрыть приложение, если пароль не введен или отменено
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
